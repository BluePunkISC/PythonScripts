from openpyxl import load_workbook
from docx import Document
 
# Rutas
ruta_excel = r'H:\Proyectos\Python\Prueba1.xlsm'
ruta_word = r'H:\Proyectos\Python\F_085_492_03_UC_nn_-_Nombre_Caso_Uso.docx'
ruta_salida = r'H:\Proyectos\Python\Nuevo_Salida_Caso_Uso.docx'
 
# Abrir Excel
wb = load_workbook(ruta_excel, data_only=True)
ws = wb.active
 
# Abrir Word
doc = Document(ruta_word)
tablas = doc.tables[1:]  # Ignorar la primera tabla (portada)
 
# Definir filas a procesar
filas_combinadas = list(range(13, 88, 5))
filas_sueltas = list(range(89, 100)) + list(range(101, 109)) + list(range(110, 124)) + list(range(137, 139))
filas_totales = filas_combinadas + filas_sueltas
 
for i, fila in enumerate(filas_totales):
    if i >= len(tablas):
        print(f"No hay suficientes tablas en el documento para la fila {fila}.")
        break
 
    tabla = tablas[i]
 
    # Leer valores del Excel
    b_val = ws[f'B{fila}'].value
    a_val = ws[f'A{fila}'].value
    f_val = ws[f'F{fila}'].value
    h_val = ws[f'H{fila}'].value
 
    # Obtener G
    if fila in filas_combinadas:
        g_partes = [ws[f'G{j}'].value for j in range(fila, fila+5)]
        g_val = ' - '.join([str(g).strip() for g in g_partes if g])
    else:
        g_val = str(ws[f'G{fila}'].value).strip() if ws[f'G{fila}'].value else ""
 
    if b_val and g_val:
        bg = f"{b_val}: {g_val}"
 
        try:
            tabla.cell(0, 1).text = bg                    # Nombre
            tabla.cell(1, 1).text = str(a_val or "")      # No Requerimiento(s) Asociado(s)
            tabla.cell(9, 1).text = str(f_val or "")      # Datos necesarios
            tabla.cell(10, 1).text = str(h_val or "")     # Descripción del proceso
        except IndexError:
            print(f"Tabla {i+1} no tiene suficientes filas o columnas.")
            continue
 
# Guardar resultado
doc.save(ruta_salida)
print("Documento Word actualizado con éxito.")